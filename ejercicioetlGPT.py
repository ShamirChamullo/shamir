import os
import pandas as pd
import re
import matplotlib.pyplot as plt
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.drawing.image import Image

def select_folder():
    folder_selected = st.text_input("Introduce la ruta de la carpeta:")
    return folder_selected

def process_files(folder, start_col, end_col, start_row):
    try:
        start_col_index = column_index_from_string(start_col.upper())
        end_col_index = column_index_from_string(end_col.upper())
        start_row = int(start_row)

        # Obtener lista de archivos Excel en la carpeta
        excel_files = [f for f in os.listdir(folder) if f.endswith('.xlsx') and f.startswith('AvanceVentasINTI')]

        all_data = []

        for i, file in enumerate(excel_files):
            file_path = os.path.join(folder, file)
            
            # Extraer año, mes y día del nombre del archivo
            match = re.search(r'AvanceVentasINTI\.(\d{4})\.(\d{2})\.(\d{2})\.', file)
            if match:
                year, month, day = match.groups()
            else:
                year, month, day = "", "", ""

            # Cargar el libro de trabajo
            wb = load_workbook(filename=file_path, read_only=True)
            sheet = wb['ITEM_O']

            # Obtener los datos de las celdas especificadas
            data = []
            for row in sheet.iter_rows(min_row=start_row, min_col=start_col_index, max_col=end_col_index):
                data.append([cell.value for cell in row])

            # Crear DataFrame
            df = pd.DataFrame(data)
            
            # Añadir columnas de año, mes y día
            df['ANIO'] = year
            df['MES'] = month
            df['DIA'] = day

            all_data.append(df)

        # Combinar todos los DataFrames
        final_df = pd.concat(all_data, ignore_index=True)

        # Generar nombres de columnas basados en las letras de las columnas
        column_names = [get_column_letter(i) for i in range(start_col_index, end_col_index + 1)]
        column_names.extend(['ANIO', 'MES', 'DIA'])
        final_df.columns = column_names

        # Exportar a Excel en la misma carpeta de entrada
        output_path = os.path.join(folder, 'Out.xlsx')
        final_df.to_excel(output_path, index=False)

        # Generar gráficos y guardarlos en una hoja aparte
        generate_charts(final_df, output_path, folder)

        st.success(f"Proceso completado. Archivo guardado en: {output_path}")
        st.dataframe(final_df.head())

    except Exception as e:
        st.error(f"Error: {str(e)}")

def generate_charts(df, output_path, folder):
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
        workbook = writer.book

        # Crear una nueva hoja para los gráficos
        charts_sheet = workbook.create_sheet(title='Charts')

        for column in df.columns:
            if df[column].dtype in ['int64', 'float64']:
                # Generar histograma
                plt.figure()
                df[column].hist()
                plt.title(f'Histograma de {column}')
                hist_path = os.path.join(folder, f'{column}_hist.png')
                plt.savefig(hist_path)
                plt.close()

                # Agregar histograma al archivo Excel
                img = Image(hist_path)
                charts_sheet.add_image(img, f'A{charts_sheet.max_row + 2}')

            # Generar gráfico de torta
            if df[column].dtype == 'object' or df[column].dtype.name == 'category':
                plt.figure()
                df[column].value_counts().plot.pie(autopct='%1.1f%%')
                plt.title(f'Torta de {column}')
                pie_path = os.path.join(folder, f'{column}_pie.png')
                plt.savefig(pie_path)
                plt.close()

                # Agregar gráfico de torta al archivo Excel
                img = Image(pie_path)
                charts_sheet.add_image(img, f'A{charts_sheet.max_row + 20}')

def main():
    st.title("Proceso ETL")

    folder = select_folder()
    start_col = st.text_input("Columna inicial (ej. A):")
    end_col = st.text_input("Columna final (ej. P):")
    start_row = st.text_input("Fila inicial:")

    if st.button("Procesar archivos"):
        process_files(folder, start_col, end_col, start_row)

if __name__ == "__main__":
    main()
